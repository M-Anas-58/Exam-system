import openpyxl
import os

ROLL_ALIASES = {"roll", "roll number", "rollnumber", "roll no", "rollno"}
NAME_ALIASES  = {"name", "names", "student name", "student names", "full name"}

def load_students_from_file(filepath):
    if not os.path.exists(filepath):
        print(f"[DATABASE] File not found: {filepath}")
        return {}

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    header_row = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in ws[1]
    ]

    roll_idx = None
    name_idx = None

    for idx, h in enumerate(header_row):
        if h in ROLL_ALIASES and roll_idx is None:
            roll_idx = idx
        if h in NAME_ALIASES and name_idx is None:
            name_idx = idx

    if roll_idx is None or name_idx is None:
        print("[DATABASE] ERROR: Could not find roll/name columns.")
        print(f"[DATABASE] Headers found: {header_row}")
        print(f"[DATABASE] Accepted roll headers : {ROLL_ALIASES}")
        print(f"[DATABASE] Accepted name headers : {NAME_ALIASES}")
        return {}

    print(f"[DATABASE] Mapped → roll col #{roll_idx}, name col #{name_idx}")

    students = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        roll_val = row[roll_idx] if roll_idx < len(row) else None
        name_val = row[name_idx] if name_idx < len(row) else None

        if not roll_val or not name_val:
            continue

        roll = str(roll_val).strip()
        name = str(name_val).strip()

        students[roll] = {
            "name":    name,
            "roll":    roll,
            "present": False
        }

    print(f"[DATABASE] Loaded {len(students)} students.")
    return students